#!/usr/bin/env python3
"""
vSphere VM Report Script (Fully Optimized)
Uses paginated PropertyCollector with pre-fetched folder hierarchy.
Reports: folder, power state, boot time, CPU/RAM/Disk usage, thin/thick provisioning.
"""

from pyVim.connect import SmartConnect, Disconnect
from pyVmomi import vim, vmodl
import ssl
import atexit
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ─── Configuration ───────────────────────────────────────────────────────────
VCENTER = "vcenter.ce.te-labs.1keyes.net"
USERNAME = "CIC user"
PASSWORD = "CIC password"
PORT = 443
# ─────────────────────────────────────────────────────────────────────────────


def connect():
    context = ssl._create_unverified_context()
    si = SmartConnect(host=VCENTER, user=USERNAME, pwd=PASSWORD, port=PORT, sslContext=context)
    atexit.register(Disconnect, si)
    print(f"[+] Connected to {VCENTER}")
    return si


def paginated_fetch(content, obj_type, container_type, props_list, batch_size=100):
    """Paginated PropertyCollector fetch."""
    container = content.viewManager.CreateContainerView(
        content.rootFolder, [obj_type], True
    )
    traversal_spec = vmodl.query.PropertyCollector.TraversalSpec(
        name='t', path='view', skip=False, type=container_type
    )
    property_spec = vmodl.query.PropertyCollector.PropertySpec(
        type=obj_type, all=False, pathSet=props_list
    )
    object_spec = vmodl.query.PropertyCollector.ObjectSpec(
        obj=container, skip=True, selectSet=[traversal_spec]
    )
    filter_spec = vmodl.query.PropertyCollector.FilterSpec(
        objectSet=[object_spec], propSet=[property_spec]
    )

    options = vmodl.query.PropertyCollector.RetrieveOptions()
    options.maxObjects = batch_size

    results = []
    partial = content.propertyCollector.RetrievePropertiesEx([filter_spec], options)
    while partial:
        results.extend(partial.objects)
        if not partial.token:
            break
        partial = content.propertyCollector.ContinueRetrievePropertiesEx(partial.token)

    container.Destroy()
    return results


def build_folder_map(content):
    """Pre-fetch all folder names and parents to avoid lazy loading."""
    results = paginated_fetch(content, vim.Folder, vim.view.ContainerView, ['name', 'parent'])

    folder_map = {}  # moref_str -> {name, parent_moref_str}
    for obj in results:
        props = {p.name: p.val for p in obj.propSet}
        moref = str(obj.obj)
        parent = props.get('parent')
        folder_map[moref] = {
            'name': props.get('name', ''),
            'parent': str(parent) if parent else ''
        }
    return folder_map


def resolve_folder_path(parent_moref_str, folder_map):
    """Resolve folder path from pre-fetched map."""
    path_parts = []
    current = parent_moref_str

    # Walk up the tree (max 10 levels to prevent infinite loops)
    for _ in range(10):
        if current not in folder_map:
            break
        info = folder_map[current]
        name = info['name']
        # Skip root VM folders and system folders
        if name in ('vm', 'Datacenters', 'host', 'datastore', 'network'):
            break
        path_parts.insert(0, name)
        current = info['parent']

    return "/".join(path_parts) if path_parts else "[Root]"


def format_size(gb):
    if gb >= 1024:
        return f"{gb/1024:.2f} TB"
    return f"{gb:.1f} GB"


def export_to_excel(vm_records):
    """Export VM data to a formatted Excel file with filters and conditional formatting."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = os.path.expanduser(f"~/vsphere_report_{timestamp}.xlsx")

    wb = Workbook()

    # ─── Sheet 1: All VMs ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "All VMs"

    # Styles
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    on_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    off_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = [
        "Folder", "VM Name", "Power State", "vCPUs", "CPU Usage (MHz)",
        "RAM Allocated (MB)", "RAM Used (MB)", "RAM Usage %",
        "Disk Capacity (GB)", "Disk Used (GB)", "Disk Usage %",
        "# Disks", "Provisioning", "Last Boot Time"
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Data rows
    for row_idx, vm in enumerate(sorted(vm_records, key=lambda x: (x['folder'], x['name'])), 2):
        # Calculate percentages
        ram_pct = (vm['mem_consumed_mb'] / vm['memory_mb'] * 100) if vm['memory_mb'] > 0 and vm['mem_consumed_mb'] > 0 else 0
        disk_pct = (vm['disk_used_gb'] / vm['disk_capacity_gb'] * 100) if vm['disk_capacity_gb'] > 0 and vm['disk_used_gb'] > 0 else 0

        row_data = [
            vm['folder'],
            vm['name'],
            vm['power_state'],
            vm['cpu_cores'],
            vm['cpu_usage_mhz'] if vm['cpu_usage_mhz'] > 0 else None,
            vm['memory_mb'],
            vm['mem_consumed_mb'] if vm['mem_consumed_mb'] > 0 else None,
            round(ram_pct, 1) if ram_pct > 0 else None,
            round(vm['disk_capacity_gb'], 1),
            round(vm['disk_used_gb'], 1) if vm['disk_used_gb'] > 0 else None,
            round(disk_pct, 1) if disk_pct > 0 else None,
            vm['num_disks'],
            vm['provisioning'],
            vm['last_event'] if vm['last_event'] != "N/A" else None,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')

        # Color rows by power state
        fill = on_fill if vm['power_state'] == "ON" else off_fill
        ws.cell(row=row_idx, column=3).fill = fill

    # Auto-fit column widths
    for col in range(1, len(headers) + 1):
        max_len = len(str(headers[col - 1]))
        for row in range(2, min(len(vm_records) + 2, 50)):  # Sample first 50 rows
            val = ws.cell(row=row, column=col).value
            if val:
                max_len = max(max_len, len(str(val)))
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 2, 35)

    # Enable auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(vm_records) + 1}"

    # Freeze top row
    ws.freeze_panes = "A2"

    # ─── Sheet 2: Summary by Folder ──────────────────────────────────────────
    ws2 = wb.create_sheet("By Folder")

    folder_headers = [
        "Folder", "Total VMs", "Powered ON", "Powered OFF",
        "Total vCPUs", "Total RAM (GB)", "Total Disk (GB)",
        "Total Disk Used (GB)"
    ]

    for col, header in enumerate(folder_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Group by folder
    folder_stats = defaultdict(lambda: {
        'total': 0, 'on': 0, 'off': 0,
        'cpu': 0, 'ram_gb': 0, 'disk_gb': 0, 'disk_used_gb': 0
    })
    for vm in vm_records:
        f = folder_stats[vm['folder']]
        f['total'] += 1
        if vm['power_state'] == "ON":
            f['on'] += 1
        else:
            f['off'] += 1
        f['cpu'] += vm['cpu_cores']
        f['ram_gb'] += vm['memory_mb'] / 1024
        f['disk_gb'] += vm['disk_capacity_gb']
        f['disk_used_gb'] += vm['disk_used_gb']

    for row_idx, (folder, stats) in enumerate(sorted(folder_stats.items()), 2):
        row_data = [
            folder, stats['total'], stats['on'], stats['off'],
            stats['cpu'], round(stats['ram_gb'], 1),
            round(stats['disk_gb'], 1), round(stats['disk_used_gb'], 1)
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    ws2.auto_filter.ref = f"A1:{get_column_letter(len(folder_headers))}{len(folder_stats) + 1}"
    ws2.freeze_panes = "A2"

    for col in range(1, len(folder_headers) + 1):
        ws2.column_dimensions[get_column_letter(col)].width = 18

    # ─── Sheet 3: Provisioning Summary ───────────────────────────────────────
    ws3 = wb.create_sheet("Provisioning")

    prov_headers = ["Provisioning Type", "Count", "Total Disk (GB)", "Total Disk Used (GB)"]
    for col, header in enumerate(prov_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    prov_stats = defaultdict(lambda: {'count': 0, 'disk_gb': 0, 'used_gb': 0})
    for vm in vm_records:
        p = prov_stats[vm['provisioning']]
        p['count'] += 1
        p['disk_gb'] += vm['disk_capacity_gb']
        p['used_gb'] += vm['disk_used_gb']

    for row_idx, (prov, stats) in enumerate(sorted(prov_stats.items()), 2):
        row_data = [prov, stats['count'], round(stats['disk_gb'], 1), round(stats['used_gb'], 1)]
        for col, value in enumerate(row_data, 1):
            cell = ws3.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border

    ws3.auto_filter.ref = f"A1:{get_column_letter(len(prov_headers))}{len(prov_stats) + 1}"
    ws3.freeze_panes = "A2"
    for col in range(1, len(prov_headers) + 1):
        ws3.column_dimensions[get_column_letter(col)].width = 22

    # ─── Save ────────────────────────────────────────────────────────────────
    wb.save(filename)
    print(f"\n[+] Excel report saved: {filename}")
    return filename


def main():
    si = connect()
    content = si.RetrieveContent()

    # ─── Pre-fetch folder hierarchy ──────────────────────────────────────────
    print("[+] Fetching folder hierarchy...")
    folder_map = build_folder_map(content)
    print(f"    Got {len(folder_map)} folders")

    # ─── Fetch VM basic properties ───────────────────────────────────────────
    print("[+] Fetching VM metadata...")
    basic_results = paginated_fetch(content, vim.VirtualMachine, vim.view.ContainerView, [
        'name', 'parent', 'runtime.powerState', 'runtime.bootTime',
        'config.hardware.numCPU', 'config.hardware.memoryMB',
        'summary.quickStats.overallCpuUsage',
        'summary.quickStats.guestMemoryUsage',
        'summary.quickStats.hostMemoryUsage',
        'summary.storage.committed',
        'summary.storage.uncommitted',
    ])
    print(f"    Got {len(basic_results)} VMs")

    # ─── Fetch disk device info ──────────────────────────────────────────────
    print("[+] Fetching disk provisioning info...")
    disk_results = paginated_fetch(content, vim.VirtualMachine, vim.view.ContainerView,
                                   ['name', 'config.hardware.device'])
    print(f"    Got {len(disk_results)} VMs")

    # Build disk info map
    disk_map = {}
    for obj in disk_results:
        props = {p.name: p.val for p in obj.propSet}
        name = props.get('name', '')
        devices = props.get('config.hardware.device', [])

        total_gb = 0
        prov_types = []
        num_disks = 0

        for d in devices:
            if isinstance(d, vim.vm.device.VirtualDisk):
                num_disks += 1
                total_gb += d.capacityInKB / (1024 * 1024)
                backing = d.backing
                if hasattr(backing, 'thinProvisioned'):
                    if backing.thinProvisioned:
                        prov_types.append("Thin")
                    else:
                        if hasattr(backing, 'eagerlyScrub') and backing.eagerlyScrub:
                            prov_types.append("Thick-Eager")
                        else:
                            prov_types.append("Thick-Lazy")
                else:
                    prov_types.append("N/A")

        unique = list(set(prov_types))
        disk_map[name] = {
            'capacity_gb': total_gb,
            'provisioning': "/".join(unique) if unique else "Unknown",
            'num_disks': num_disks,
        }

    # ─── Build VM records ────────────────────────────────────────────────────
    print("[+] Building report...\n")
    vms_by_folder = defaultdict(list)

    for obj in basic_results:
        props = {p.name: p.val for p in obj.propSet}

        name = props.get('name', 'Unknown')
        power_state = str(props.get('runtime.powerState', ''))
        is_on = 'poweredOn' in power_state
        state_str = "ON" if is_on else "OFF"

        cpu_cores = props.get('config.hardware.numCPU', 0)
        memory_mb = props.get('config.hardware.memoryMB', 0)

        boot_time = props.get('runtime.bootTime')
        if boot_time and is_on:
            last_event = f"ON since {boot_time.strftime('%Y-%m-%d %H:%M')}"
        else:
            last_event = "N/A"

        cpu_usage = (props.get('summary.quickStats.overallCpuUsage') or 0) if is_on else 0
        mem_consumed = (props.get('summary.quickStats.hostMemoryUsage') or 0) if is_on else 0

        committed = props.get('summary.storage.committed') or 0
        disk_used_gb = committed / (1024**3)

        dinfo = disk_map.get(name, {'capacity_gb': 0, 'provisioning': 'Unknown', 'num_disks': 0})

        # Resolve folder path from pre-fetched map (no lazy loading)
        parent_obj = props.get('parent')
        parent_moref_str = str(parent_obj) if parent_obj else ''
        folder_name = resolve_folder_path(parent_moref_str, folder_map)

        vms_by_folder[folder_name].append({
            "name": name,
            "power_state": state_str,
            "cpu_cores": cpu_cores,
            "memory_mb": memory_mb,
            "cpu_usage_mhz": cpu_usage,
            "mem_consumed_mb": mem_consumed,
            "disk_capacity_gb": dinfo['capacity_gb'],
            "disk_used_gb": disk_used_gb,
            "num_disks": dinfo['num_disks'],
            "provisioning": dinfo['provisioning'],
            "last_event": last_event,
        })

    # ─── Build flat list for Excel ──────────────────────────────────────────
    all_vm_records = []
    for folder_name in sorted(vms_by_folder.keys()):
        for vm in vms_by_folder[folder_name]:
            all_vm_records.append({**vm, "folder": folder_name})

    # ─── Export to Excel ─────────────────────────────────────────────────────
    export_to_excel(all_vm_records)

    # ─── Print Summary ───────────────────────────────────────────────────────
    total_vms = len(all_vm_records)
    total_on = sum(1 for v in all_vm_records if v["power_state"] == "ON")
    total_off = total_vms - total_on
    total_cpu = sum(v["cpu_cores"] for v in all_vm_records)
    total_mem_mb = sum(v["memory_mb"] for v in all_vm_records)
    total_disk_gb = sum(v["disk_capacity_gb"] for v in all_vm_records)

    print(f"\n{'=' * 80}")
    print(f" SUMMARY")
    print(f"{'=' * 80}")
    print(f" Total VMs:              {total_vms}")
    print(f" Powered ON:             {total_on}")
    print(f" Powered OFF:            {total_off}")
    print(f" Total Folders:          {len(vms_by_folder)}")
    print(f" Total vCPUs Allocated:  {total_cpu}")
    print(f" Total RAM Allocated:    {format_size(total_mem_mb / 1024)}")
    print(f" Total Disk Provisioned: {format_size(total_disk_gb)}")
    print(f"{'=' * 80}")
    print("\n[+] Done.")


if __name__ == "__main__":
    main()
